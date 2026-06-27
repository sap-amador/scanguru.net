#!/usr/bin/env python3
"""
add_language.py — one-command pipeline to add/complete a ScanGuru i18n language.

The English source is FIXED: wired_en.json (1336 ordered unique values).
Every language is just those 1336 slots translated. Supply a payload and this does
the rest: glossary -> build_packs --refresh -> engine validate -> full-site verify
-> Excel column -> clinical_review_pending.json. No per-language scoping needed.

USAGE
  python3 add_language.py <lang> <payload.json> [--no-deploy-files]

PAYLOAD (either form):
  * dict:  { "<english value>": "<target>", ... }   (order-independent, robust)
  * list:  [ "<target>", ... ] length 1336, positional against wired_en.json
           (or assemble from the 4 chunk arrays in order)

Run from repo root. Outputs glossary_<lang>.json, the clinical sheet, and the tracker to repo root.
"""
import json, os, sys, subprocess, shutil

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = HERE  # root layout: script lives in repo root
ENGINE = os.path.join(REPO, "scanguru-i18n-v6-full-site.js")
OUTDIR = REPO  # root layout: outputs land in repo root
SHEET = os.path.join(OUTDIR, "scanguru_clinical_terms.xlsx")
CRP = os.path.join(OUTDIR, "clinical_review_pending.json")

def die(m): print("ERROR:", m); sys.exit(1)

def load_canon():
    return json.load(open(os.path.join(HERE, "wired_en.json"), encoding="utf-8"))

def engine_flatten_to_json(lang, outpath):
    """Use node to flatten engine translations[lang] -> {dotpath: value} JSON."""
    js = f'''
const I=require({json.dumps(ENGINE)});
const fs=require("fs");
const o=I.translations[{json.dumps(lang)}]||{{}};
const m={{}};(function w(x,p){{for(const k in x){{const v=x[k];const np=p?p+"."+k:k;
 if(v&&typeof v==="object")w(v,np);else m[np]=v;}}}})(o,"");
fs.writeFileSync({json.dumps(outpath)},JSON.stringify(m));
'''
    p = os.path.join(HERE, "_flatten_tmp.js")
    open(p, "w").write(js)
    subprocess.run(["node", p], check=True)
    os.remove(p)
    return json.load(open(outpath, encoding="utf-8"))

def main():
    if len(sys.argv) < 3:
        die("usage: add_language.py <lang> <payload.json> [--no-deploy-files]")
    lang = sys.argv[1]
    payload_path = sys.argv[2]
    canon = load_canon()
    payload = json.load(open(payload_path, encoding="utf-8"))

    # ---- build {en_value: target} map ----
    if isinstance(payload, dict):
        vmap = payload
    elif isinstance(payload, list):
        if len(payload) != len(canon):
            die(f"positional payload length {len(payload)} != canonical {len(canon)}")
        vmap = dict(zip(canon, payload))
    else:
        die("payload must be a JSON object or array")

    missing = [v for v in canon if v not in vmap or vmap[v] in (None, "")]
    if missing:
        die(f"{len(missing)} canonical values have no translation. First few: {missing[:5]}")

    # ---- write glossary_<lang>.json (seed from engine first, then overlay payload) ----
    os.chdir(REPO)
    seed_flat = engine_flatten_to_json(lang, os.path.join(HERE, "_seedflat.json"))
    os.remove(os.path.join(HERE, "_seedflat.json"))
    en_flat = engine_flatten_to_json("en", os.path.join(HERE, "_enflat.json"))
    os.remove(os.path.join(HERE, "_enflat.json"))
    glossary = {}
    for k, ev in en_flat.items():            # carry pre-existing engine coverage as seed
        if k in seed_flat and ev is not None:
            glossary[ev] = seed_flat[k]
    glossary.update(vmap)                     # payload wins
    gpath = os.path.join(REPO, f"glossary_{lang}.json")
    json.dump(glossary, open(gpath, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    os.makedirs(OUTDIR, exist_ok=True)
    _dst = os.path.join(OUTDIR, f"glossary_{lang}.json")
    if os.path.abspath(_dst) != os.path.abspath(gpath):   # root layout: gpath IS already in OUTDIR
        shutil.copy(gpath, _dst)
    print(f"glossary_{lang}.json: {len(glossary)} pairs")

    # ---- regenerate en_leaves.json then build_packs --refresh ----
    js = f'''const fs=require("fs");const I=require({json.dumps(ENGINE)});const out=[];
(function w(o,p){{for(const k in o){{const v=o[k];const np=p?p+"."+k:k;
 if(v&&typeof v==="object")w(v,np);else if(typeof v==="string")out.push([np,v]);}}}})(I.translations.en,"");
fs.writeFileSync("en_leaves.json",JSON.stringify(out));'''
    p = os.path.join(REPO, "_enleaves_tmp.js"); open(p, "w").write(js)
    subprocess.run(["node", p], check=True); os.remove(p)
    r = subprocess.run(["python3", "build_packs.py", lang, f"glossary_{lang}.json", "--refresh"],
                       capture_output=True, text=True)
    print("build_packs:", (r.stdout.strip().splitlines() or ["(no output)"])[-1])
    subprocess.run(["node", "-e", f"require({json.dumps(ENGINE)});console.log('engine OK')"], check=True)
    print("OK — engine valid. Run verify_all.js next, or it runs automatically below.")

    # ---- verify (delegate to verify_all.js for all engine languages) ----
    subprocess.run(["node", os.path.join(HERE, "verify_all.js")], check=True)

    # ---- sheet + clinical tracker (best-effort; skip if files absent) ----
    try:
        update_sheet_and_tracker(lang)
    except Exception as e:
        print("sheet/tracker update skipped:", e)

def update_sheet_and_tracker(lang):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    lf = engine_flatten_to_json(lang, os.path.join(HERE, "_lf.json"))
    os.remove(os.path.join(HERE, "_lf.json"))

    if os.path.exists(SHEET):
        wb = load_workbook(SHEET); ws = wb["Clinical Terms"]
        col = ws.max_column + 1
        # if this lang already has a column, reuse it
        for c in range(5, ws.max_column + 1):
            if (ws.cell(1, c).value or "").startswith(f"{lang} "):
                col = c; break
        ws.cell(1, col).value = f"{lang} (live)"
        ws.cell(1, col).fill = PatternFill("solid", start_color="0F766E")
        ws.cell(1, col).font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        ws.cell(1, col).alignment = Alignment(vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
        filled = 0
        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 2).value
            cell = ws.cell(r, col)
            cell.value = lf.get(key, "") if key else ""
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True); cell.border = border
            if ws.cell(r, 3).value == "scoring":
                cell.fill = PatternFill("solid", start_color="F1F5F9")
            if cell.value: filled += 1
        ws.column_dimensions[get_column_letter(col)].width = 30
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        wb.save(SHEET)
        print(f"sheet: {lang} column filled {filled} rows (col {get_column_letter(col)})")

    if os.path.exists(CRP):
        rp = json.load(open(CRP, encoding="utf-8"))
        n = 0
        for it in rp.get("items", []):
            k = it.get("key"); it[f"{lang}_live"] = lf.get(k, "") if k else ""
            if it[f"{lang}_live"]: n += 1
        langs = rp.get("languages_live_unreviewed", [])
        if lang not in langs: langs.append(lang)
        rp["languages_live_unreviewed"] = langs
        json.dump(rp, open(CRP, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"clinical_review_pending: {lang}_live added to {n} rows; langs now {langs}")

if __name__ == "__main__":
    main()
